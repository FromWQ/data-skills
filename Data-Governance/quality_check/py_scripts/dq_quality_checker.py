#!/usr/bin/env python3
"""
数仓数据质量评估执行脚本
一键执行完整评估流程：初始化、同步元数据、执行校验、生成报告
"""

import argparse
import os
import sys
from datetime import datetime
from typing import List, Dict

# 添加 py_scripts 目录到路径，以便导入 db_connector
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from db_connector import DbConnector

# Excel 导出相关导入（可选）
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ============ 配置常量 ============
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)

INPUT_DIR = os.path.join(BASE_DIR, 'input')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')
REFERENCES_DIR = os.path.join(BASE_DIR, 'references')

DB_CONFIG_FILE = os.path.join(INPUT_DIR, 'db_config.json')
RULE_FILE = os.path.join(REFERENCES_DIR, 'rule_v2.md')

os.makedirs(OUTPUT_DIR, exist_ok=True)

LAYER_WEIGHTS = {'ODS': 0.15, 'DWD': 0.25, 'DWS': 0.25, 'DIM': 0.20, 'ADS': 0.15}
SEVERITY_SCORES = {'BLOCK': 25, 'WARN': 10, 'INFO': 3}
SCORE_LEVELS = [(90, 100, '优秀'), (75, 89, '良好'), (60, 74, '一般'), (40, 59, '较差'), (0, 39, '危险')]


def get_score_level(score):
    """根据评分返回等级"""
    if score >= 90:
        return '优秀'
    elif score >= 75:
        return '良好'
    elif score >= 60:
        return '一般'
    elif score >= 40:
        return '较差'
    else:
        return '危险'


def init_environment():
    """初始化环境：创建表（阶段 0）"""
    print("\n" + "=" * 60)
    print("🔧 阶段 0：环境初始化...")
    print("=" * 60)

    db = DbConnector()
    try:
        # 检查 meta_table_column 表是否存在
        result = db.query_data("""
            SELECT TABLE_NAME FROM information_schema.TABLES
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'meta_table_column'
        """, (db.config.get('database'),))

        if not result:
            # 表不存在，创建表
            db.execute_sql("""
                CREATE TABLE IF NOT EXISTS `meta_table_column` (
                    `id`                BIGINT          NOT NULL AUTO_INCREMENT COMMENT '主键 ID',
                    `table_name`        VARCHAR(128)    NOT NULL COMMENT '表名称',
                    `table_comment`     VARCHAR(512)    DEFAULT NULL COMMENT '表注释/描述',
                    `table_layer`       VARCHAR(16)     NOT NULL COMMENT '表所属分层：ODS/DWD/DWS/DIM/ADS',
                    `column_name`       VARCHAR(64)     NOT NULL COMMENT '字段名称',
                    `column_type`       VARCHAR(64)     NOT NULL COMMENT '字段类型',
                    `column_comment`    VARCHAR(512)    DEFAULT NULL COMMENT '字段注释/描述',
                    `column_tag`        VARCHAR(16)     NOT NULL DEFAULT 'ATTRIBUTE' COMMENT '字段标签',
                    `is_pk`             TINYINT(1)      DEFAULT 0 COMMENT '是否主键',
                    `is_partition`      TINYINT(1)      DEFAULT 0 COMMENT '是否分区字段',
                    `created_at`        DATETIME        NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    `updated_at`        DATETIME        NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    PRIMARY KEY (`id`),
                    UNIQUE KEY `uk_table_column` (`table_name`, `column_name`),
                    KEY `idx_table_name` (`table_name`),
                    KEY `idx_table_layer` (`table_layer`)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='表字段元数据表'
            """)
            print("✅ 创建 meta_table_column 表")
        else:
            print("✅ meta_table_column 表已存在")
            # 表存在，检查是否需要添加 table_layer 字段
            result = db.query_data("""
                SELECT COLUMN_NAME FROM information_schema.COLUMNS
                WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'meta_table_column' AND COLUMN_NAME = 'table_layer'
            """, (db.config.get('database'),))

            if not result:
                db.execute_sql("ALTER TABLE meta_table_column ADD COLUMN table_layer VARCHAR(16) AFTER table_comment")
                print("✅ 添加 table_layer 字段")
            else:
                print("✅ table_layer 字段已存在")

        result = db.query_data("""
            SELECT TABLE_NAME FROM information_schema.TABLES
            WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'dq_check_result'
        """, (db.config.get('database'),))

        if not result:
            # 表不存在，创建完整结构的表
            db.execute_sql("""
                CREATE TABLE IF NOT EXISTS `dq_check_result` (
                    `id`                    BIGINT          NOT NULL AUTO_INCREMENT PRIMARY KEY,
                    `table_name`            VARCHAR(128)    NOT NULL,
                    `table_layer`           VARCHAR(16)     NOT NULL,
                    `partition_value`       VARCHAR(32)     DEFAULT NULL,
                    `total_rows`            BIGINT          DEFAULT 0,
                    `rule_id`               VARCHAR(16)     NOT NULL,
                    `rule_name`             VARCHAR(64)     NOT NULL,
                    `rule_severity`         VARCHAR(16)     NOT NULL,
                    `check_field`           VARCHAR(128)    DEFAULT NULL,
                    `check_sql`             TEXT            NOT NULL,
                    `execute_status`        VARCHAR(16)     NOT NULL,
                    `execute_time`          INT             DEFAULT 0,
                    `error_message`         TEXT            DEFAULT NULL,
                    `check_status`          VARCHAR(16)     NOT NULL,
                    `anomaly_count`         BIGINT          DEFAULT 0,
                    `anomaly_rate`          DECIMAL(10,4)   DEFAULT 0.0000,
                    `check_standard`        VARCHAR(256)    DEFAULT NULL,
                    `actual_value`          TEXT            DEFAULT NULL,
                    `analysis_result`       TEXT            DEFAULT NULL,
                    `check_time`            DATETIME        DEFAULT CURRENT_TIMESTAMP,
                    KEY `idx_table_name` (`table_name`),
                    KEY `idx_table_layer` (`table_layer`),
                    KEY `idx_rule_id` (`rule_id`),
                    KEY `idx_check_status` (`check_status`)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
            print("✅ 创建 dq_check_result 表")
        else:
            print("✅ dq_check_result 表已存在")
            # 表已存在，检查并添加缺失字段
            required_fields = [
                ('partition_value', 'VARCHAR(32)', 'NULL'),
                ('error_message', 'TEXT', 'NULL'),
                ('anomaly_rate', 'DECIMAL(10,4)', 'DEFAULT 0.0000'),
                ('check_standard', 'VARCHAR(256)', 'NULL'),
                ('actual_value', 'TEXT', 'NULL'),
                ('analysis_result', 'TEXT', 'NULL')
            ]

            existing_fields = db.query_data("""
                SELECT COLUMN_NAME FROM information_schema.COLUMNS
                WHERE TABLE_SCHEMA = %s AND TABLE_NAME = 'dq_check_result'
            """, (db.config.get('database'),))
            existing_field_names = [row['COLUMN_NAME'] for row in existing_fields]

            for field_name, field_type, field_extra in required_fields:
                if field_name not in existing_field_names:
                    db.execute_sql(f"ALTER TABLE dq_check_result ADD COLUMN `{field_name}` {field_type} {field_extra}")
                    print(f"✅ 添加 {field_name} 字段到 dq_check_result 表")

        for layer, prefix in [('ODS', 'ods%'), ('DWD', 'dwd%'), ('DWS', 'dws%'), ('DIM', 'dim%'), ('ADS', 'ads%')]:
            db.execute_sql(f"UPDATE meta_table_column SET table_layer = '{layer}' WHERE table_name LIKE '{prefix}' AND (table_layer IS NULL OR table_layer = '')")

        result = db.query_data("SELECT table_layer, COUNT(DISTINCT table_name) as cnt FROM meta_table_column WHERE table_layer IS NOT NULL GROUP BY table_layer")
        print("\n📊 各分层表数量:")
        for row in result:
            print(f"   {row['table_layer']}: {row['cnt']} 张表")

        print("\n✅ 环境初始化完成!")
    finally:
        db.disconnect()


def sync_metadata():
    """同步元数据（阶段 1）"""
    print("\n" + "=" * 60)
    print("🔄 阶段 1：同步元数据...")
    print("=" * 60)
    db = DbConnector()
    try:
        db.sync_table_metadata()
    finally:
        db.disconnect()


def load_rules(rule_file: str) -> List[Dict]:
    """加载校验规则"""
    rules = []
    try:
        with open(rule_file, 'r', encoding='utf-8') as f:
            content = f.read()

        # 简单解析规则文件
        current_rule = {}
        for line in content.split('\n'):
            line = line.strip()
            if line.startswith('### '):
                if current_rule:
                    rules.append(current_rule)
                rule_id = line.replace('### ', '').split()[0]
                current_rule = {'rule_id': rule_id, 'rule_name': line.replace(f'### {rule_id} ', '')}
            elif line.startswith('**严重程度**'):
                current_rule['severity'] = line.split('**')[2]
            elif line.startswith('**SQL 模板**'):
                current_rule['sql_template'] = True
            elif line.startswith('**校验标准**'):
                current_rule['standard'] = line.split('**')[2]

        if current_rule:
            rules.append(current_rule)

        print(f"✅ 加载 {len(rules)} 条规则")
    except Exception as e:
        print(f"⚠️  规则文件加载失败：{e}")

    return rules


def get_table_fields(db, table_name: str) -> Dict:
    """获取表的字段信息"""
    sql = f"""
        SELECT column_name, column_type, column_comment, is_pk, is_partition
        FROM meta_table_column
        WHERE table_name = '{table_name}'
    """
    result = db.query_data(sql)

    fields = {
        'all': [],
        'pk': [],
        'partition': [],
        'time': [],
        'measure': [],
        'varchar': []
    }

    time_keywords = ['time', 'date', 'datetime', 'timestamp', '时间', '日期']
    measure_keywords = ['mny', 'amount', 'price', 'cost', 'fee', 'qty', 'count', 'rate', '金额', '价格', '成本', '费用', '数量']

    for row in result:
        col_name = row['column_name']
        col_type = row['column_type'].lower()
        col_comment = (row['column_comment'] or '').lower()

        fields['all'].append(col_name)

        if row['is_pk'] == 1:
            fields['pk'].append(col_name)
        if row['is_partition'] == 1:
            fields['partition'].append(col_name)
        if any(kw in col_name.lower() or kw in col_comment for kw in time_keywords) or 'time' in col_type or 'date' in col_type:
            fields['time'].append(col_name)
        if any(kw in col_name.lower() or kw in col_comment for kw in measure_keywords):
            fields['measure'].append(col_name)
        if 'varchar' in col_type or 'text' in col_type:
            fields['varchar'].append(col_name)

    return fields


def run_checks():
    """执行质量检查（阶段 2）"""
    print("\n" + "=" * 60)
    print("🔍 阶段 2：执行质量检查...")
    print("=" * 60)

    db = DbConnector()
    try:
        # 获取表清单
        tables = db.get_table_list()
        print(f"📊 发现 {len(tables)} 个数仓表")

        # 加载规则
        rules = load_rules(RULE_FILE)

        # 获取元数据表中的表信息
        table_metadata = db.query_data("""
            SELECT DISTINCT table_name, table_layer
            FROM meta_table_column
            ORDER BY table_layer, table_name
        """)

        total_checks = 0
        passed_checks = 0
        failed_checks = 0

        for table_info in table_metadata:
            table_name = table_info['table_name']
            table_layer = table_info['table_layer']

            print(f"\n   📋 检查表：{table_name} ({table_layer})")

            # 获取表记录数
            count_sql = f"SELECT COUNT(*) as cnt FROM {table_name}"
            try:
                count_result = db.query_data(count_sql)
                total_rows = count_result[0]['cnt'] if count_result else 0
            except:
                total_rows = 0

            # 获取字段信息
            fields = get_table_fields(db, table_name)

            # 根据字段情况应用规则
            checks_for_table = []

            # COM_003: 关键字段非空率 (如果有主键字段)
            if fields['pk']:
                key_field = fields['pk'][0]
                sql = f"""
                    SELECT
                        COUNT(*) as total_count,
                        SUM(CASE WHEN {key_field} IS NULL OR {key_field} = '' THEN 1 ELSE 0 END) as null_count
                    FROM {table_name}
                """
                try:
                    result = db.query_data(sql)
                    total_count = result[0]['total_count'] or 0
                    null_count = result[0]['null_count'] or 0

                    check_status = 'PASS' if null_count == 0 else 'FAIL'
                    total_checks += 1

                    if check_status == 'PASS':
                        passed_checks += 1
                    else:
                        failed_checks += 1

                    checks_for_table.append({
                        'rule_id': 'COM_003',
                        'rule_name': '关键字段非空检查',
                        'severity': 'BLOCK',
                        'field': key_field,
                        'status': check_status,
                        'anomaly_count': null_count,
                        'total_rows': total_rows
                    })

                    print(f"      COM_003: {check_status} (空值：{null_count}/{total_count})")
                except Exception as e:
                    print(f"      COM_003: 执行失败 - {e}")

            # COM_004: 时间字段无异常 (如果有时间字段)
            if fields['time']:
                time_field = fields['time'][0]
                sql = f"""
                    SELECT
                        COUNT(*) as total_count,
                        SUM(CASE WHEN {time_field} < '1970-01-01' OR {time_field} > NOW() THEN 1 ELSE 0 END) as invalid_count
                    FROM {table_name}
                """
                try:
                    result = db.query_data(sql)
                    total_count = result[0]['total_count'] or 0
                    invalid_count = result[0]['invalid_count'] or 0

                    check_status = 'PASS' if invalid_count == 0 else 'FAIL'
                    total_checks += 1

                    if check_status == 'PASS':
                        passed_checks += 1
                    else:
                        failed_checks += 1

                    checks_for_table.append({
                        'rule_id': 'COM_004',
                        'rule_name': '时间字段合理性检查',
                        'severity': 'WARN',
                        'field': time_field,
                        'status': check_status,
                        'anomaly_count': invalid_count,
                        'total_rows': total_rows
                    })

                    print(f"      COM_004: {check_status} (异常：{invalid_count}/{total_count})")
                except Exception as e:
                    print(f"      COM_004: 执行失败 - {e}")

            # 写入检查结果到数据库
            for check in checks_for_table:
                insert_sql = f"""
                    INSERT INTO dq_check_result (
                        table_name, table_layer, total_rows,
                        rule_id, rule_name, rule_severity, check_field,
                        check_sql, execute_status, execute_time,
                        check_status, anomaly_count, analysis_result, check_time
                    ) VALUES (
                        '{table_name}', '{table_layer}', {check['total_rows']},
                        '{check['rule_id']}', '{check['rule_name']}', '{check['severity']}', '{check['field']}',
                        '', 'SUCCESS', 0,
                        '{check['status']}', {check['anomaly_count']},
                        '异常数：{check["anomaly_count"]}',
                        NOW()
                    )
                """
                try:
                    db.execute_sql(insert_sql)
                except Exception as e:
                    print(f"        写入结果失败：{e}")

        print(f"\n✅ 检查完成！共执行 {total_checks} 项检查，通过 {passed_checks} 项，失败 {failed_checks} 项")

    finally:
        db.disconnect()


def generate_report():
    """生成评估报告（阶段 3）"""
    print("\n" + "=" * 60)
    print("📝 阶段 3：生成评估报告...")
    print("=" * 60)

    db = DbConnector()
    try:
        # 获取检查结果
        results = db.query_data("""
            SELECT table_name, table_layer, rule_id, rule_name, rule_severity,
                   check_status, anomaly_count, total_rows, check_field, analysis_result
            FROM dq_check_result
            ORDER BY table_layer, table_name, rule_id
        """)

        if not results:
            print("⚠️  没有找到检查结果，请先执行质量检查")
            return

        # 计算统计信息
        total_rules = len(results)
        passed_rules = sum(1 for r in results if r['check_status'] == 'PASS')
        failed_rules = total_rules - passed_rules
        pass_rate = (passed_rules / total_rules * 100) if total_rules > 0 else 0

        block_failures = sum(1 for r in results if r['rule_severity'] == 'BLOCK' and r['check_status'] == 'FAIL')
        warn_failures = sum(1 for r in results if r['rule_severity'] == 'WARN' and r['check_status'] == 'FAIL')
        info_failures = sum(1 for r in results if r['rule_severity'] == 'INFO' and r['check_status'] == 'FAIL')

        # 计算单表得分
        table_scores = {}
        for r in results:
            table = r['table_name']
            if table not in table_scores:
                table_scores[table] = {'layer': r['table_layer'], 'score': 100, 'rules': []}

            if r['check_status'] == 'FAIL':
                if r['rule_severity'] == 'BLOCK':
                    table_scores[table]['score'] -= 25
                elif r['rule_severity'] == 'WARN':
                    table_scores[table]['score'] -= 10
                elif r['rule_severity'] == 'INFO':
                    table_scores[table]['score'] -= 3

            table_scores[table]['rules'].append(r)
            table_scores[table]['score'] = max(0, table_scores[table]['score'])  # 最低 0 分

        # 计算分层得分
        layer_scores = {}
        for table, data in table_scores.items():
            layer = data['layer']
            if layer not in layer_scores:
                layer_scores[layer] = {'total': 0, 'count': 0}
            layer_scores[layer]['total'] += data['score']
            layer_scores[layer]['count'] += 1

        for layer in layer_scores:
            if layer_scores[layer]['count'] > 0:
                layer_scores[layer]['avg'] = layer_scores[layer]['total'] / layer_scores[layer]['count']
            else:
                layer_scores[layer]['avg'] = 0

        # 计算整体得分（带权重）
        weights = {'ODS': 0.15, 'DWD': 0.25, 'DWS': 0.25, 'DIM': 0.20, 'ADS': 0.15}
        overall_score = sum(layer_scores.get(layer, {}).get('avg', 0) * weight
                           for layer, weight in weights.items())

        # 生成报告内容
        report_path = os.path.join(OUTPUT_DIR, 'quality_check_report.md')

        # 问题规则统计
        rule_failures = {}
        for r in results:
            if r['check_status'] == 'FAIL':
                key = (r['rule_id'], r['rule_name'], r['rule_severity'])
                if key not in rule_failures:
                    rule_failures[key] = 0
                rule_failures[key] += 1

        rule_failures_sorted = sorted(rule_failures.items(), key=lambda x: x[1], reverse=True)[:10]

        # 各分层低分表排行
        layer_tables = {'ODS': [], 'DWD': [], 'DWS': [], 'DIM': [], 'ADS': []}
        for table, data in table_scores.items():
            layer = data['layer']
            if layer in layer_tables:
                layer_tables[layer].append((table, data['score']))

        for layer in layer_tables:
            layer_tables[layer].sort(key=lambda x: x[1])  # 低分在前

        report_content = f"""# 数仓数据质量评估报告

**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
**评估版本**: v2.0

---

## 开篇总结

本次评估共检查 **{total_rules}** 条规则，通过 **{passed_rules}** 条，通过率 **{pass_rate:.1f}%**。

**整体健康评分**: {overall_score:.2f} 分（{get_score_level(overall_score)}）

**核心问题**:
- 阻断级 (BLOCK) 问题：{block_failures} 个
- 警告级 (WARN) 问题：{warn_failures} 个
- 提示级 (INFO) 问题：{info_failures} 个

---

## 一、评估概览

### 1.1 整体评分

| 指标 | 值 |
|------|-----|
| 整体评分 | {overall_score:.2f} 分 ({get_score_level(overall_score)}) |
| 规则总数 | {total_rules} |
| 通过规则 | {passed_rules} |
| 规则通过率 | {pass_rate:.1f}% |
| 阻断问题数 | {block_failures} |
| 警告问题数 | {warn_failures} |
| 提示问题数 | {info_failures} |

### 1.2 分层评分

| 分层 | 评分 | 权重 | 加权分 |
|------|------|------|--------|
"""
        # 添加分层评分
        for layer in ['ODS', 'DWD', 'DWS', 'DIM', 'ADS']:
            if layer in layer_scores and layer_scores[layer]['count'] > 0:
                score = layer_scores[layer]['avg']
                weight = weights.get(layer, 0)
                weighted = score * weight
                report_content += f"| {layer} | {score:.2f} | {weight:.1%} | {weighted:.1f} |\n"

        report_content += f"""
### 1.3 评分等级说明

| 分数范围 | 等级 | 说明 |
|---------|------|------|
| 90-100 | 优秀 | |
| 75-89 | 良好 | |
| 60-74 | 一般 | |
| 40-59 | 较差 | |
| 0-39 | 危险 | |

---

## 二、问题分布

### 2.1 严重程度分布

| 严重程度 | 问题数 | 占比 |
|---------|--------|------|
| BLOCK (阻断) | {block_failures} | {block_failures/max(1,failed_rules)*100:.1f}% |
| WARN (警告) | {warn_failures} | {warn_failures/max(1,failed_rules)*100:.1f}% |
| INFO (提示) | {info_failures} | {info_failures/max(1,failed_rules)*100:.1f}% |

### 2.2 问题规则 TOP10

| 排名 | 规则 ID | 规则名称 | 严重程度 | 失败次数 |
|------|---------|---------|---------|--------|
"""
        for i, ((rule_id, rule_name, severity), count) in enumerate(rule_failures_sorted, 1):
            report_content += f"| {i} | {rule_id} | {rule_name} | {severity} | {count} |\n"

        if not rule_failures_sorted:
            report_content += "| - | 无问题 | - | - | 0 |\n"

        report_content += f"""
---

## 三、分层详情

### 3.1 各分层表评分排行

"""
        for layer in ['ODS', 'DWD', 'DWS', 'DIM', 'ADS']:
            tables = layer_tables.get(layer, [])
            if tables:
                report_content += f"\n#### {layer}层 TOP5 (低分)\n\n"
                report_content += "| 排名 | 表名 | 评分 |\n|------|------|------|\n"
                for i, (table, score) in enumerate(tables[:5], 1):
                    report_content += f"| {i} | {table} | {score} |\n"

        failed_results = [r for r in results if r['check_status'] == 'FAIL']

        report_content += f"""
---

## 四、问题处理建议

### 4.1 阻断级问题清单 (BLOCK)

| 表名 | 规则 | 问题描述 | 处理建议 |
|------|------|---------|--------|
"""
        block_failed = [r for r in failed_results if r['rule_severity'] == 'BLOCK']
        if block_failed:
            for r in block_failed:
                report_content += f"| {r['table_name']} | {r['rule_name']} | {r['rule_name']} | 立即检查上游数据源和 ETL 逻辑 |\n"
        else:
            report_content += "| - | 无 | - | - |\n"

        report_content += f"""
### 4.2 警告级问题清单 (WARN)

| 表名 | 规则 | 问题描述 |
|------|------|---------|
"""
        warn_failed = [r for r in failed_results if r['rule_severity'] == 'WARN']
        if warn_failed:
            for r in warn_failed:
                report_content += f"| {r['table_name']} | {r['rule_name']} | {r['rule_name']} |\n"
        else:
            report_content += "| - | 无 | - |\n"

        report_content += f"""
---

## 五、趋势对比

> 暂无历史数据，首次评估

---

## 六、附录

### 6.1 评估规则清单

本次评估共应用 **{len(set(r['rule_id'] for r in results))}** 条规则。

### 6.2 数据来源说明

- 元数据来源：`meta_table_column` 表
- 校验结果：`dq_check_result` 表
- 规则定义：`rule_v2.md`

---

**报告生成完成** ✅
"""

        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(report_content)

        print(f"\n✅ 评估报告已保存：{report_path}")
        print(f"   整体评分：{overall_score:.2f} 分（{get_score_level(overall_score)}）")
        print(f"   规则通过率：{pass_rate:.1f}%")

        # 导出 Excel 结果表
        export_excel_result()

    finally:
        db.disconnect()


def export_excel_result():
    """从 dq_check_result 表导出数据到 Excel"""
    print("\n  导出 Excel 结果表...")

    if not OPENPYXL_AVAILABLE:
        print("  ⚠️  openpyxl 库未安装，跳过 Excel 导出")
        print("     安装命令：pip3 install openpyxl")
        return

    db = DbConnector()
    try:
        # 从 dq_check_result 表读取数据
        results = db.query_data("""
            SELECT table_name, table_layer, rule_id, rule_name, rule_severity,
                   check_field, check_status, anomaly_count, total_rows, anomaly_rate,
                   check_standard, actual_value, analysis_result, check_time
            FROM dq_check_result
            ORDER BY table_layer, table_name, rule_id
        """)

        if not results:
            print("  ⚠️  没有检查结果数据")
            return

        excel_path = os.path.join(OUTPUT_DIR, 'quality_check_result.xlsx')

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = '质量检查结果'

        # 表头
        headers = ['表名', '分层', '规则 ID', '规则名称', '严重程度', '检查字段',
                   '检查状态', '异常数量', '总记录数', '异常率 (%)',
                   '校验标准', '实际值', '分析结果', '检查时间']

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 数据行
        for row, result in enumerate(results, start=2):
            ws.cell(row=row, column=1, value=result['table_name'])
            ws.cell(row=row, column=2, value=result['table_layer'])
            ws.cell(row=row, column=3, value=result['rule_id'])
            ws.cell(row=row, column=4, value=result['rule_name'])
            ws.cell(row=row, column=5, value=result['rule_severity'])
            ws.cell(row=row, column=6, value=result['check_field'] or '-')

            status_cell = ws.cell(row=row, column=7, value=result['check_status'])
            if result['check_status'] == 'PASS':
                status_cell.font = Font(color='00B050')
            else:
                status_cell.font = Font(color='FF0000')

            ws.cell(row=row, column=8, value=result['anomaly_count'])
            ws.cell(row=row, column=9, value=result['total_rows'])

            # 格式化异常率
            anomaly_rate = result['anomaly_rate']
            if anomaly_rate:
                ws.cell(row=row, column=10, value=float(anomaly_rate))
            else:
                ws.cell(row=row, column=10, value=0)

            ws.cell(row=row, column=11, value=result['check_standard'] or '-')
            ws.cell(row=row, column=12, value=result['actual_value'] or '-')
            ws.cell(row=row, column=13, value=result['analysis_result'] or '-')

            check_time = result['check_time']
            if check_time:
                ws.cell(row=row, column=14, value=check_time.strftime('%Y-%m-%d %H:%M:%S'))
            else:
                ws.cell(row=row, column=14, value='-')

        # 设置列宽
        column_widths = [20, 10, 12, 25, 12, 15, 12, 12, 12, 12, 20, 20, 30, 20]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # 设置第一行加粗
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 保存
        wb.save(excel_path)
        print(f"  ✅ Excel 结果表已保存：{excel_path}")
        print(f"     共 {len(results)} 条记录")

    finally:
        db.disconnect()


def main():
    """主函数"""
    parser = argparse.ArgumentParser(description='数仓数据质量评估工具')
    parser.add_argument('--init', action='store_true', help='初始化环境（创建表）')
    parser.add_argument('--sync', action='store_true', help='同步元数据')
    parser.add_argument('--check', action='store_true', help='执行质量检查')
    parser.add_argument('--report', action='store_true', help='生成评估报告')
    parser.add_argument('--all', action='store_true', help='执行完整流程')

    args = parser.parse_args()

    if args.all:
        init_environment()
        sync_metadata()
        run_checks()
        generate_report()
    else:
        if args.init:
            init_environment()
        if args.sync:
            sync_metadata()
        if args.check:
            run_checks()
        if args.report:
            generate_report()

    if not any([args.init, args.sync, args.check, args.report, args.all]):
        parser.print_help()


if __name__ == '__main__':
    main()
